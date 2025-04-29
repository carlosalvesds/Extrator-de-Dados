
import streamlit as st
import pandas as pd
import zipfile
import tempfile
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Extrator de Contas de Energia", layout="centered")

st.title("⚡ Extrator de Contas de Energia Elétrica (PDF ➜ Excel)")

uploaded_files = st.file_uploader("Envie arquivos PDF ou um arquivo .ZIP com vários PDFs:", type=["pdf", "zip"], accept_multiple_files=True)

def criar_planilha(dados_pdf):
    wb = Workbook()
    for uc, info in dados_pdf.items():
        ws = wb.create_sheet(title=uc)

        # Cabeçalho principal
        ws.append(["Número da Nota Fiscal", info["Número da Nota Fiscal"]])
        ws.append(["Data de Emissão", info["Data de Emissão"]])
        ws.append(["CNPJ", info["CNPJ"]])
        ws.append(["Nome do Titular", info["Nome do Titular"]])
        ws.append(["Valor Total NF", info["Valor Total NF"]])
        ws.append([])

        # Dados de fornecimento
        df = pd.DataFrame(info["Fornecimento"])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Formatação
        for row in ws.iter_rows(min_row=6, max_row=6, max_col=3):
            for cell in row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="404040")
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=7, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(cell.value, (int, float)):
                    cell.number_format = 'R$ #,##0.00'

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max_length + 2

        ws.sheet_view.showGridLines = False

    # Remover a aba padrão
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def processar_arquivos(uploaded_files):
    arquivos_pdf = []

    for uploaded in uploaded_files:
        if uploaded.type == "application/zip":
            with tempfile.TemporaryDirectory() as tmpdir:
                zip_path = os.path.join(tmpdir, "temp.zip")
                with open(zip_path, "wb") as f:
                    f.write(uploaded.read())
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(tmpdir)
                    for file_name in zip_ref.namelist():
                        if file_name.lower().endswith(".pdf"):
                            with open(os.path.join(tmpdir, file_name), "rb") as f:
                                arquivos_pdf.append((file_name, f.read()))
        else:
            arquivos_pdf.append((uploaded.name, uploaded.read()))
    return arquivos_pdf

if uploaded_files:
    arquivos_pdf = processar_arquivos(uploaded_files)

    st.warning("⚠️ Esta versão demo utiliza dados simulados extraídos manualmente.")

    dados_demo = {
        "760064611": {
            "Número da Nota Fiscal": "142599856",
            "Data de Emissão": "15/04/2025",
            "CNPJ": "03.465.317/0001-91",
            "Nome do Titular": "ROMA EMPREENDIMENTOS E TURISMO LTDA",
            "Valor Total NF": 78432.55,
            "Fornecimento": [
                {"Fornecimento": "CONSUMO NÃO COMPENSADO FP - TUSD kWh", "Valor (R$)": 4551.38, "ICMS (R$)": 864.76},
                {"Fornecimento": "CONSUMO NÃO COMPENSADO HR - TUSD kWh", "Valor (R$)": 7359.28, "ICMS (R$)": 1398.26},
                {"Fornecimento": "CONSUMO NÃO COMPENSADO P - TUSD kWh", "Valor (R$)": 3031.61, "ICMS (R$)": 576.01},
                {"Fornecimento": "INJEÇÃO SCEE FP - TE - GD I kWh", "Valor (R$)": -18.83, "ICMS (R$)": -3.58},
                {"Fornecimento": "INJEÇÃO SCEE FP - TUSD - GD I kWh", "Valor (R$)": -83.72, "ICMS (R$)": -15.91},
            ]
        },
        "10029306637": {
            "Número da Nota Fiscal": "142599507",
            "Data de Emissão": "15/04/2025",
            "CNPJ": "12.947.899/0001-33",
            "Nome do Titular": "JARDINS DI ROMA COMERCIAL LTDA",
            "Valor Total NF": 11145.88,
            "Fornecimento": [
                {"Fornecimento": "CONSUMO FP SCEE - TUSD kWh", "Valor (R$)": 7563.37, "ICMS (R$)": 1437.04},
                {"Fornecimento": "CONSUMO HR SCEE - TUSD kWh", "Valor (R$)": 544.84, "ICMS (R$)": 103.52},
                {"Fornecimento": "INJEÇÃO SCEE FP - TUSD - UC ...", "Valor (R$)": -420.45, "ICMS (R$)": -79.89},
                {"Fornecimento": "INJEÇÃO SCEE HR - TUSD - UC ...", "Valor (R$)": -544.84, "ICMS (R$)": -103.52},
            ]
        },
        "10038540701": {
            "Número da Nota Fiscal": "140836989",
            "Data de Emissão": "03/04/2025",
            "CNPJ": "03.465.317/0001-91",
            "Nome do Titular": "ROMA EMPREENDIMENTOS E TURISMO LTDA",
            "Valor Total NF": 692.59,
            "Fornecimento": [
                {"Fornecimento": "CONSUMO kWh kWh", "Valor (R$)": 660.61, "ICMS (R$)": 125.52}
            ]
        }
    }

    excel_bytes = criar_planilha(dados_demo)
    st.success("Planilha gerada com sucesso!")
    st.download_button("📥 Baixar Excel", data=excel_bytes, file_name="Extrato_Contas_Energia.xlsx")
else:
    st.info("Envie os arquivos PDF ou um ZIP para iniciar.")
